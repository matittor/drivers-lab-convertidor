import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import filedialog, messagebox, Tk
import os

def seleccionar_archivos():
    root = Tk()
    root.withdraw()
    archivos = filedialog.askopenfilenames(
        title="Selecciona archivos de medición",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    return list(archivos)

def seleccionar_plantilla():
    root = Tk()
    root.withdraw()
    plantilla = filedialog.askopenfilename(
        title="Selecciona la plantilla base",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    return plantilla

def seleccionar_salida():
    root = Tk()
    root.withdraw()
    salida = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos Excel", "*.xlsx")],
        title="Guardar archivo como..."
    )
    return salida

def extraer_datos_validos(filepath):
    xls = pd.ExcelFile(filepath)
    for sheet in xls.sheet_names:
        df = xls.parse(sheet, header=None)
        for i, row in df.iterrows():
            if row.astype(str).str.contains("V out", na=False).any():
                df_clean = xls.parse(sheet, header=i)
                voltaje = "120V" if "120" in filepath else "277V"
                df_clean["Voltaje"] = voltaje
                return df_clean
    return pd.DataFrame()

def insertar_datos_en_plantilla(plantilla_path, datos_120, datos_277, salida_path):
    wb = openpyxl.load_workbook(plantilla_path)
    ws = wb.active

    # Buscar secciones
    fila_120 = fila_277 = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and any("Test @120Vac" in str(cell) for cell in row if cell):
            fila_120 = i + 2
        if row and any("Test @277Vac" in str(cell) for cell in row if cell):
            fila_277 = i + 2

    if datos_120 is not None and fila_120:
        for r in dataframe_to_rows(datos_120, index=False, header=True):
            for idx, val in enumerate(r):
                ws.cell(row=fila_120, column=idx+1, value=val)
            fila_120 += 1

    if datos_277 is not None and fila_277:
        for r in dataframe_to_rows(datos_277, index=False, header=True):
            for idx, val in enumerate(r):
                ws.cell(row=fila_277, column=idx+1, value=val)
            fila_277 += 1

    wb.save(salida_path)
    messagebox.showinfo("Éxito", f"Archivo guardado exitosamente:\n{salida_path}")

# --- EJECUCIÓN ---
if __name__ == "__main__":
    archivos = seleccionar_archivos()
    if not archivos:
        messagebox.showwarning("Cancelado", "No se seleccionaron archivos.")
        exit()

    plantilla = seleccionar_plantilla()
    if not plantilla:
        messagebox.showwarning("Cancelado", "No se seleccionó plantilla.")
        exit()

    salida = seleccionar_salida()
    if not salida:
        messagebox.showwarning("Cancelado", "No se seleccionó archivo de salida.")
        exit()

    datos_120 = datos_277 = None
    for archivo in archivos:
        datos = extraer_datos_validos(archivo)
        if "120" in archivo:
            datos_120 = datos
        elif "277" in archivo:
            datos_277 = datos

    insertar_datos_en_plantilla(plantilla, datos_120, datos_277, salida)
